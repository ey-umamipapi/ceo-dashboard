"""
sync_masterpapi.py
------------------
Reads MasterPapi_FY26_LIVE.xlsx and MarginPapi - Chilli Oil.xlsx,
then upserts all data into Supabase.

Tables written:
  revenue_monthly, revenue_weekly  — from MasterPapi Mth/Wk sheets
  prod_runs                        — per-day production data from PROD sheet
  prod_efficiency                  — monthly capacity metrics from Production Dashboard
  margin_skus                      — Chilli Oil margin by product/channel from MarginPapi

Usage:
    python scripts/sync_masterpapi.py

Requirements:
    pip install openpyxl supabase python-dotenv

Supabase tables required (run once in SQL editor):
    CREATE TABLE prod_runs (
      id SERIAL PRIMARY KEY,
      run_date DATE NOT NULL UNIQUE,
      staff NUMERIC, hours_worked NUMERIC, comments TEXT,
      sku1 TEXT, sku1_tins INTEGER,
      sku2 TEXT, sku2_tins INTEGER,
      sku3 TEXT, sku3_tins INTEGER,
      sku4 TEXT, sku4_tins INTEGER,
      total_tins INTEGER, tins_per_hour NUMERIC
    );
    CREATE TABLE prod_efficiency (
      id SERIAL PRIMARY KEY,
      month TEXT NOT NULL UNIQUE,
      sort_order INTEGER,
      demand INTEGER, capacity NUMERIC, surplus NUMERIC,
      utilisation NUMERIC, staff_required NUMERIC, current_staff NUMERIC
    );
    CREATE TABLE margin_skus (
      id SERIAL PRIMARY KEY,
      product TEXT NOT NULL, channel TEXT NOT NULL,
      sell_price NUMERIC, cogs NUMERIC,
      margin_dollars NUMERIC, margin_pct NUMERIC, volume INTEGER, sort_order INTEGER,
      UNIQUE(product, channel)
    );
"""

import os
import sys
import shutil
import tempfile
from pathlib import Path
from datetime import datetime, date, timedelta
from dotenv import load_dotenv
import openpyxl

# ── Config ───────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
load_dotenv(SCRIPT_DIR.parent / '.env.secret')

SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_KEY')

HOME = Path.home()
ONEDRIVE_NAME = "OneDrive - umamipapi.com.au"
ONEDRIVE_MAC  = "Library/CloudStorage/OneDrive-umamipapi.com.au"

WINDOWS_ONEDRIVE = HOME / ONEDRIVE_NAME
MAC_ONEDRIVE     = HOME / ONEDRIVE_MAC

if WINDOWS_ONEDRIVE.exists():
    ONEDRIVE = WINDOWS_ONEDRIVE
elif MAC_ONEDRIVE.exists():
    ONEDRIVE = MAC_ONEDRIVE
else:
    print(f"ERROR: OneDrive not found at {WINDOWS_ONEDRIVE} or {MAC_ONEDRIVE}")
    sys.exit(1)

MASTER_PAPI_PATH = ONEDRIVE / "DatabasePapi" / "Cowork" / "Inputs" / "MasterPapi_FY26_LIVE2.xlsx"

COSTING_DIR = ONEDRIVE / "DatabasePapi" / "Costing"

# All active production SKUs. NPD files (Scallion Ginger, Japanese BBQ) live in M2.
# data_issue=True flags rows in Supabase so HubPapi shows a warning — don't suppress the data.
MARGIN_PAPI_FILES = [
    {
        "path":       COSTING_DIR / "MarginPapi - Chilli Oil (Apr 26).xlsx",
        "data_issue": False,
        "data_issue_note": None,
    },
    {
        "path":       COSTING_DIR / "MarginPapi - Hot Honey (Apr 26).xlsx",
        "data_issue": False,
        "data_issue_note": None,
    },
    {
        "path":       COSTING_DIR / "MarginPapi - Nandos PSE (Apr 26).xlsx",
        "data_issue": True,
        "data_issue_note": "COGS double-count: HELA PCRK base included in PSE cost. Fix needed by Richard. Margin understated.",
    },
    {
        "path":       COSTING_DIR / "MarginPapi - Nandos PCRK (Apr 26).xlsx",
        "data_issue": True,
        "data_issue_note": "COGS double-count: HELA PSE base included in PCRK cost. Fix needed by Richard. Margin understated.",
    },
    {
        "path":       COSTING_DIR / "MarginPapi - Mayo (Apr26).xlsx",
        "data_issue": False,
        "data_issue_note": None,
    },
]

if not MASTER_PAPI_PATH.exists():
    print(f"ERROR: MasterPapi not found at {MASTER_PAPI_PATH}")
    sys.exit(1)

# FY26 month number → short name
MONTH_MAP = {
    1: 'Jul', 2: 'Aug', 3: 'Sep', 4: 'Oct', 5: 'Nov', 6: 'Dec',
    7: 'Jan', 8: 'Feb', 9: 'Mar', 10: 'Apr', 11: 'May', 12: 'Jun'
}

CALENDAR_TO_FY = {
    7: 1, 8: 2, 9: 3, 10: 4, 11: 5, 12: 6,
    1: 7, 2: 8, 3: 9, 4: 10, 5: 11, 6: 12
}

# ── Helpers ──────────────────────────────────────────────────────────────────

def val(cell):
    v = cell.value if hasattr(cell, 'value') else cell
    if v is None:
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0

def current_fy_month():
    return CALENDAR_TO_FY[datetime.now().month]

def time_to_hours(t):
    """Convert a datetime.time to decimal hours. Returns 0 if None or midnight."""
    if not t or not hasattr(t, 'hour'):
        return 0.0
    return t.hour + t.minute / 60.0

def open_workbook(path):
    """Open a workbook, copying to temp first to handle file locks."""
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    shutil.copy2(str(path), tmp.name)
    return openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)

# ── Revenue readers ───────────────────────────────────────────────────────────

def read_monthly(wb):
    ws = wb['Mth']
    current_fy_mo = current_fy_month()
    records, sort_order = [], 0

    for row in ws.iter_rows(min_row=2):
        date_cell = row[8].value
        fy_cell   = row[9].value
        if date_cell is None or fy_cell != 'FY26':
            continue

        cal_month = None
        if hasattr(date_cell, 'month'):
            cal_month = date_cell.month
        else:
            date_str = str(date_cell).strip().lower()
            for i, m in enumerate(['january','february','march','april','may','june',
                                   'july','august','september','october','november','december'], 1):
                if m in date_str:
                    cal_month = i
                    break

        if cal_month is None:
            continue

        fy_month_num = CALENDAR_TO_FY.get(cal_month)
        if fy_month_num is None:
            continue

        total = round(val(row[50]))
        if total == 0:
            continue

        sort_order += 1
        month_name = MONTH_MAP[fy_month_num]
        is_mtd = (fy_month_num == current_fy_mo)

        record = {
            'fiscal_year': 'fy26',
            'month':   month_name,
            'total':   total,
            'direct':  round(val(row[41])),
            'wsale':   round(val(row[42])),
            'distrbn': round(val(row[43])),
            'coles':   round(val(row[44])),
            'metcash': round(val(row[45])),
            'fserv':   round(val(row[46])),
            'nandos':  round(val(row[47])),
            'other':   round(val(row[48])),
            'orders':  round(val(row[39])),
            'ad': 0, 'roas': 0,
            'mtd': is_mtd,
            'sort_order': sort_order,
        }
        records.append(record)
        print(f"  Monthly FY26 {month_name}: ${record['total']:,}{' ← MTD' if is_mtd else ''}")

    return records


def read_weekly(wb):
    ws = wb['Wk']
    records, sort_order = [], 0

    for row in ws.iter_rows(min_row=2):
        if row[9].value != 'FY26':
            continue
        date_cell = row[8].value
        if date_cell is None:
            continue
        total = round(val(row[44]))
        if total == 0:
            continue

        week_label = date_cell.strftime('%d %b').lstrip('0') if hasattr(date_cell, 'strftime') else str(date_cell)
        sort_order += 1
        records.append({
            'week_label': week_label,
            'total':   total,
            'direct':  round(val(row[35])),
            'coles':   round(val(row[38])),
            'distrbn': round(val(row[37])),
            'nandos':  round(val(row[41])),
            'other':   round(val(row[42])),
            'sort_order': sort_order,
        })

    print(f"  Weekly: {len(records)} weeks found")
    return records

# ── Production readers ────────────────────────────────────────────────────────

def read_prod_runs(wb):
    """
    PROD sheet column mapping (0-indexed):
      M=12  Date
      N=13  Staff count
      O=14  Comments
      R=17  Hours worked (time object)
      S=18  SKU #1 name   AH=33  SKU #1 net tins
      T=19  SKU #2 name   AI=34  SKU #2 net tins
      U=20  SKU #3 name   AJ=35  SKU #3 net tins
      V=21  SKU #4 name   AK=36  SKU #4 net tins
      AR=43 Jarring efficiency (min/staff-hr)
    Only rows with staff > 0 and hours > 0 are real production days.
    """
    ws = wb['PROD']
    runs = []

    for row in ws.iter_rows(min_row=19, max_row=300, values_only=True):
        date_val = row[12] if len(row) > 12 else None
        staff    = row[13] if len(row) > 13 else None
        hours_t  = row[17] if len(row) > 17 else None
        comments = row[14] if len(row) > 14 else None

        if not date_val or not hasattr(date_val, 'strftime'):
            continue
        if not staff or float(staff) == 0:
            continue
        hours_dec = time_to_hours(hours_t)
        if hours_dec == 0:
            continue

        sku1 = row[18] if len(row) > 18 else None
        sku2 = row[19] if len(row) > 19 else None
        sku3 = row[20] if len(row) > 20 else None
        sku4 = row[21] if len(row) > 21 else None

        t1 = int(row[33]) if len(row) > 33 and row[33] else 0  # AH
        t2 = int(row[34]) if len(row) > 34 and row[34] else 0  # AI
        t3 = int(row[35]) if len(row) > 35 and row[35] else 0  # AJ
        t4 = int(row[36]) if len(row) > 36 and row[36] else 0  # AK
        total = t1 + t2 + t3 + t4
        tph   = round(total / hours_dec, 2) if hours_dec > 0 else 0

        jarring_eff = row[43] if len(row) > 43 and row[43] else None
        if jarring_eff is not None:
            jarring_eff = round(float(jarring_eff), 4)

        runs.append({
            'run_date':           date_val.strftime('%Y-%m-%d'),
            'staff':              float(staff),
            'hours_worked':       round(hours_dec, 2),
            'comments':           str(comments)[:500] if comments else None,
            'sku1':               str(sku1) if sku1 else None,
            'sku1_tins':          t1 or None,
            'sku2':               str(sku2) if sku2 else None,
            'sku2_tins':          t2 or None,
            'sku3':               str(sku3) if sku3 else None,
            'sku3_tins':          t3 or None,
            'sku4':               str(sku4) if sku4 else None,
            'sku4_tins':          t4 or None,
            'total_tins':         total,
            'tins_per_hour':      tph,
            'jarring_efficiency': jarring_eff,
        })

    print(f"  Prod runs: {len(runs)} active production days")
    return runs


def read_prod_efficiency(wb):
    """
    Production Summary sheet — actual monthly tins by channel.
    Row 3:  month headers (col 1=July … col 12=June, 0-indexed)
    Row 4:  Coles
    Row 5:  Distrbn
    Row 6:  Nandos
    Row 7:  Fserv
    Row 8:  Wsale
    Row 9:  Direct
    Row 10: GRAND TOTAL
    """
    sheet_name = 'Production Summary'
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠ Sheet '{sheet_name}' not found — skipping prod_efficiency.")
        return []
    ws = wb[sheet_name]

    rows = {}
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=10, values_only=True), start=3):
        rows[i] = list(row)

    month_labels = rows.get(3, [])
    coles_row    = rows.get(4, [])
    distrbn_row  = rows.get(5, [])
    nandos_row   = rows.get(6, [])
    fserv_row    = rows.get(7, [])
    wsale_row    = rows.get(8, [])
    direct_row   = rows.get(9, [])
    total_row    = rows.get(10, [])

    # FY month order: col 1=Jul(1), col 2=Aug(2), ... col 12=Jun(12)
    FY_MONTHS = ['Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']

    records = []
    for col_idx in range(1, 13):  # cols 1–12 (July–June)
        total = total_row[col_idx] if len(total_row) > col_idx else None
        if not total:
            continue

        def safe_int(row, idx):
            v = row[idx] if len(row) > idx else None
            return int(v) if v else 0

        records.append({
            'month':        FY_MONTHS[col_idx - 1],
            'sort_order':   col_idx,
            'total_tins':   safe_int(total_row, col_idx),
            'coles_tins':   safe_int(coles_row, col_idx),
            'distrbn_tins': safe_int(distrbn_row, col_idx),
            'nandos_tins':  safe_int(nandos_row, col_idx),
            'fserv_tins':   safe_int(fserv_row, col_idx),
            'wsale_tins':   safe_int(wsale_row, col_idx),
            'direct_tins':  safe_int(direct_row, col_idx),
        })

    print(f"  Prod efficiency: {len(records)} months (from Production Summary)")
    return records

# ── Margin reader ─────────────────────────────────────────────────────────────

def read_margin_skus_from_file(file_cfg):
    """
    Reads a single MarginPapi file → MARGIN sheet.
    Column indices (0-based, consistent across all active SKU files):
      E=4   sort order
      F=5   product name (variant)
      G=6   channel name
      R=17  sell price (Apr 2026)
      AE=30 COGS (Apr 2026)
      AR=43 margin $ (Apr 2026)
      BE=56 margin % (Apr 2026, decimal e.g. 0.828)
      BF=57 volume
      BI=60 show/hide filter
    Only rows with BI='show' and non-zero price are included.
    """
    path = Path(file_cfg["path"])
    if not path.exists():
        print(f"  ⚠ Not found, skipping: {path.name}")
        return []

    file_modified_at = datetime.fromtimestamp(path.stat().st_mtime).date().isoformat()

    wb = open_workbook(path)

    # Setup!F17 is the Richard sign-off date cell (standard across all MarginPapi files).
    # Falls back to file_modified_at until Richard starts entering dates.
    last_reviewed_at = None
    try:
        ws_setup = wb['Setup']
        setup_rows = list(ws_setup.iter_rows(min_row=17, max_row=17, values_only=True))
        if setup_rows:
            reviewed_val = setup_rows[0][5]  # col F (0-indexed)
            if reviewed_val and str(reviewed_val).strip() not in ('--', '—', '-', 'Not in use', ''):
                if hasattr(reviewed_val, 'strftime'):
                    last_reviewed_at = reviewed_val.strftime('%Y-%m-%d')
                else:
                    last_reviewed_at = str(reviewed_val).strip()
    except Exception:
        pass
    if not last_reviewed_at:
        last_reviewed_at = file_modified_at

    ws = wb['MARGIN']

    PRICE_IDX  = 17
    COGS_IDX   = 30
    MARG_D_IDX = 43
    MARG_P_IDX = 56
    VOL_IDX    = 57
    BI_IDX     = 60

    records, sort_order = [], 0
    for row in ws.iter_rows(min_row=10, max_row=300, values_only=True):
        if len(row) <= BI_IDX:
            continue
        if row[BI_IDX] != 'show':
            continue

        product = row[5] if len(row) > 5 else None
        channel = row[6] if len(row) > 6 else None
        if not product or not channel:
            continue

        channel_clean = str(channel).replace('\n', ' ').strip()
        if channel_clean in ('--', '—', '-', 'Channel'):
            continue

        price = row[PRICE_IDX] if len(row) > PRICE_IDX else None
        if not price or float(price) == 0:
            continue

        cogs   = float(row[COGS_IDX]   or 0) if len(row) > COGS_IDX   else 0
        marg_d = float(row[MARG_D_IDX] or 0) if len(row) > MARG_D_IDX else 0
        marg_p = float(row[MARG_P_IDX] or 0) if len(row) > MARG_P_IDX else 0
        volume = int(row[VOL_IDX]) if len(row) > VOL_IDX and row[VOL_IDX] else 0

        sort_order += 1
        records.append({
            'product':          str(product),
            'channel':          channel_clean,
            'sell_price':       round(float(price), 2),
            'cogs':             round(cogs, 4),
            'margin_dollars':   round(marg_d, 4),
            'margin_pct':       round(marg_p, 4),
            'volume':           volume,
            'sort_order':       sort_order,
            'source_file':      path.name,
            'file_modified_at': file_modified_at,
            'last_reviewed_at': last_reviewed_at,
            'data_issue':       file_cfg["data_issue"],
            'data_issue_note':  file_cfg["data_issue_note"],
            'synced_at':        datetime.now().__str__(),
        })

    wb.close()
    flag = " ⚠ DATA ISSUE" if file_cfg["data_issue"] else ""
    print(f"  {path.stem}: {len(records)} rows, modified {file_modified_at}{flag}")
    return records


def read_all_margin_skus():
    """Loop all active MarginPapi files and return combined records."""
    print(f"  Reading {len(MARGIN_PAPI_FILES)} MarginPapi files...")
    all_records = []
    for file_cfg in MARGIN_PAPI_FILES:
        all_records.extend(read_margin_skus_from_file(file_cfg))
    print(f"  Total margin rows: {len(all_records)}")
    return all_records

# ── Inventory reader ─────────────────────────────────────────────────────────

def read_inventory():
    """
    Returns inventory snapshot records using the hardcoded fallback values
    from Operations.tsx. When the PROD sheet or a live inventory source is
    wired up, replace this with a real reader.
    """
    today = datetime.now().strftime('%Y-%m-%d')
    hardcoded = [
        {'sku': 'OG Large',        'available': 4210,  'status': 'OK'},
        {'sku': 'ES Large',        'available': 152,   'status': 'critical'},
        {'sku': 'ES Jumbo',        'available': 70,    'status': 'critical'},
        {'sku': 'OG Jumbo',        'available': 55,    'status': 'critical'},
        {'sku': 'Chilli Egg Mayo', 'available': 17185, 'status': 'watch'},
        {'sku': 'Hot Honey',       'available': 2100,  'status': 'OK'},
        {'sku': 'PERi Crackle 1KG','available': 890,   'status': 'OK'},
    ]
    records = [
        {**row, 'snapshot_date': today}
        for row in hardcoded
    ]
    print(f"  Inventory: {len(records)} SKUs (hardcoded stub, date={today})")
    return records

def read_daily_metrics(wb):
    """
    Read daily production metrics from Metrics_draft sheet.
    Columns: E=week_label, F=commentary, G=date, K=tins_produced,
             L=tins_filled_overnight, M=tins_filled_day,
             O=jars_filled_overnight, P=jars_filled_day
    """
    ws = wb['Metrics_draft']
    records = []
    now = datetime.utcnow().isoformat()
    current_week  = None
    current_comment = None

    for row in ws.iter_rows(min_row=12, max_row=ws.max_row, values_only=True):
        if len(row) < 12:
            continue

        week_cell    = row[4]  # col E
        comment_cell = row[5]  # col F
        date_val     = row[6]  # col G
        tins_prod    = row[10] # col K
        tins_night   = row[11] # col L
        tins_day     = row[12] if len(row) > 12 else None  # col M
        jars_night   = row[14] if len(row) > 14 else None  # col O
        jars_day     = row[15] if len(row) > 15 else None  # col P

        if week_cell and isinstance(week_cell, str) and week_cell.startswith('WEEK'):
            current_week = week_cell.strip()
        if comment_cell and isinstance(comment_cell, str) and comment_cell.strip():
            current_comment = comment_cell.strip()[:500]

        if not date_val or not hasattr(date_val, 'strftime'):
            continue

        d = date_val.date() if hasattr(date_val, 'date') else date_val

        def safe_int(v):
            try:
                return int(float(v)) if v is not None else None
            except (TypeError, ValueError):
                return None

        records.append({
            'metric_date':          d.strftime('%Y-%m-%d'),
            'week_label':           current_week,
            'commentary':           current_comment if week_cell and isinstance(week_cell, str) and week_cell.startswith('WEEK') else None,
            'tins_produced':        safe_int(tins_prod),
            'tins_filled_overnight': safe_int(tins_night),
            'tins_filled_day':      safe_int(tins_day),
            'jars_filled_overnight': safe_int(jars_night),
            'jars_filled_day':      safe_int(jars_day),
            'synced_at':            now,
        })

    print(f"  Daily metrics: {len(records)} rows")
    return records


def read_inventory_batches(wb):
    """
    Read active batch codes from INV sheet BATCH LOCATER.
    Columns: F=OG_LRG, G=OG_JBO, H=OG_OTH, I=ES_LRG, J=ES_JBO, K=ES_OTH, L=CEM, M=OTH
    """
    import re
    ws = wb['INV']
    now = datetime.utcnow().isoformat()
    records = []
    # Header row is row 10 (index 9), data starts row 11
    PRODUCT_COLS = {
        5:  'OG_LRG',
        6:  'OG_JBO',
        7:  'OG_OTH',
        8:  'ES_LRG',
        9:  'ES_JBO',
        10: 'ES_OTH',
        11: 'CEM',
        12: 'OTH',
    }

    def parse_expiry(code):
        if not code or not isinstance(code, str):
            return None
        m = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})$', code)
        if not m:
            return None
        d, mo, y = m.groups()
        y = int(y)
        if y < 100:
            y += 2000
        try:
            return f"{y:04d}-{int(mo):02d}-{int(d):02d}"
        except Exception:
            return None

    seen = set()
    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True):
        for col_idx, product_type in PRODUCT_COLS.items():
            val = row[col_idx] if len(row) > col_idx else None
            if not val or not isinstance(val, str):
                continue
            code = val.strip()
            if not code or code in seen or code.startswith('#'):
                continue
            seen.add(code)
            records.append({
                'batch_code':   code,
                'product_type': product_type,
                'expiry_date':  parse_expiry(code),
                'synced_at':    now,
            })

    print(f"  Inventory batches: {len(records)} batch codes")
    return records


def read_prod_schedule(wb):
    """
    Read upcoming production schedule from PROD sheet.
    Returns rows for the next 21 days where SKUs are planned,
    regardless of whether hours/staff are filled in yet.
    """
    ws   = wb['PROD']
    today      = date.today()
    look_ahead = today + timedelta(days=21)
    schedule   = []

    DAY_NAMES = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    for row in ws.iter_rows(min_row=19, max_row=400, values_only=True):
        date_val = row[12] if len(row) > 12 else None
        if not date_val or not hasattr(date_val, 'strftime'):
            continue
        d = date_val.date() if hasattr(date_val, 'date') else date_val
        if d < today or d > look_ahead:
            continue

        sku1 = row[18] if len(row) > 18 else None
        sku2 = row[19] if len(row) > 19 else None
        sku3 = row[20] if len(row) > 20 else None
        sku4 = row[21] if len(row) > 21 else None
        skus = [s for s in [sku1, sku2, sku3, sku4] if s]
        if not skus:
            continue  # skip days with nothing planned

        staff    = row[13] if len(row) > 13 else None
        hours_t  = row[17] if len(row) > 17 else None
        comments = row[14] if len(row) > 14 else None
        hours_dec = time_to_hours(hours_t)

        schedule.append({
            'run_date':   d.strftime('%Y-%m-%d'),
            'day_name':   DAY_NAMES[d.weekday()],
            'sku1':       str(sku1) if sku1 else None,
            'sku2':       str(sku2) if sku2 else None,
            'sku3':       str(sku3) if sku3 else None,
            'sku4':       str(sku4) if sku4 else None,
            'staff':      int(float(staff)) if staff else None,
            'hours':      round(hours_dec, 2) if hours_dec > 0 else None,
            'notes':      str(comments)[:200] if comments else None,
        })

    print(f"  Prod schedule: {len(schedule)} upcoming days")
    return schedule


# ── Read MasterPapi ───────────────────────────────────────────────────────────

def read_masterpapi(path):
    path = Path(path)
    print(f"Opening: {path.name}")
    wb = open_workbook(path)

    monthly           = read_monthly(wb)
    weekly            = read_weekly(wb)
    prod_runs         = read_prod_runs(wb)
    prod_efficiency   = read_prod_efficiency(wb)
    prod_schedule     = read_prod_schedule(wb)
    daily_metrics     = read_daily_metrics(wb)
    inventory_batches = read_inventory_batches(wb)

    wb.close()
    return monthly, weekly, prod_runs, prod_efficiency, prod_schedule, daily_metrics, inventory_batches

# ── Upsert to Supabase ────────────────────────────────────────────────────────

def upsert_supabase(monthly, weekly, prod_runs, prod_efficiency, prod_schedule, margin_skus, inventory_snap, daily_metrics, inventory_batches):
    from supabase import create_client

    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERROR: SUPABASE_URL or SUPABASE_SERVICE_KEY missing from .env.secret")
        sys.exit(1)

    print(f"\nConnecting to Supabase: {SUPABASE_URL}")
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    # Revenue monthly
    if monthly:
        print(f"\nUpserting {len(monthly)} monthly revenue rows...")
        sb.table('revenue_monthly').delete().eq('fiscal_year', 'fy26').execute()
        sb.table('revenue_monthly').insert(monthly).execute()
        print(f"✓ revenue_monthly: {len(monthly)} rows")

    # Revenue weekly
    if weekly:
        print(f"Upserting {len(weekly)} weekly revenue rows...")
        sb.table('revenue_weekly').delete().neq('id', 0).execute()
        sb.table('revenue_weekly').insert(weekly).execute()
        print(f"✓ revenue_weekly: {len(weekly)} rows")

    # Production runs
    if prod_runs:
        print(f"Upserting {len(prod_runs)} production runs...")
        sb.table('prod_runs').delete().neq('id', 0).execute()
        sb.table('prod_runs').insert(prod_runs).execute()
        print(f"✓ prod_runs: {len(prod_runs)} rows")

    # Production efficiency
    if prod_efficiency:
        print(f"Upserting {len(prod_efficiency)} efficiency months...")
        sb.table('prod_efficiency').delete().neq('id', 0).execute()
        sb.table('prod_efficiency').insert(prod_efficiency).execute()
        print(f"✓ prod_efficiency: {len(prod_efficiency)} rows")

    # Margin SKUs — full replace on every sync (single source of truth: MarginPapi files)
    if margin_skus:
        print(f"Upserting {len(margin_skus)} margin SKU rows across {len(MARGIN_PAPI_FILES)} files...")
        sb.table('margin_skus').delete().neq('id', 0).execute()
        sb.table('margin_skus').insert(margin_skus).execute()
        print(f"✓ margin_skus: {len(margin_skus)} rows")

    # Production schedule
    if prod_schedule:
        print(f"Upserting {len(prod_schedule)} prod schedule rows...")
        sb.table('prod_schedule').delete().neq('run_date', '').execute()
        sb.table('prod_schedule').insert(prod_schedule).execute()
        print(f"✓ prod_schedule: {len(prod_schedule)} rows")

    # Inventory snapshot
    if inventory_snap:
        print(f"Upserting {len(inventory_snap)} inventory snapshot rows...")
        for row in inventory_snap:
            sb.table('inventory_snapshot').upsert(row, on_conflict='sku,snapshot_date').execute()
        print(f"✓ inventory_snapshot: {len(inventory_snap)} rows")

    # Daily metrics
    if daily_metrics:
        print(f"Upserting {len(daily_metrics)} daily metric rows...")
        sb.table('daily_metrics').delete().neq('id', 0).execute()
        sb.table('daily_metrics').insert(daily_metrics).execute()
        print(f"✓ daily_metrics: {len(daily_metrics)} rows")

    # Inventory batches
    if inventory_batches:
        print(f"Upserting {len(inventory_batches)} inventory batch codes...")
        sb.table('inventory_batches').delete().neq('id', 0).execute()
        sb.table('inventory_batches').insert(inventory_batches).execute()
        print(f"✓ inventory_batches: {len(inventory_batches)} rows")

    # Sync timestamps — must be UTC ISO so the browser displays the correct local date
    from datetime import timezone
    now = datetime.now(timezone.utc).isoformat()
    for source in ('masterpapi', 'financial', 'seo', 'marketing'):
        sb.table('sync_metadata').update({'last_sync_at': now}).eq('source', source).execute()
    print(f"✓ sync_metadata: all timestamps updated to {now}")

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("UmamiPapi — MasterPapi → Supabase Sync")
    print(f"Run at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    print("\n[1/3] Reading MasterPapi...")
    monthly, weekly, prod_runs, prod_efficiency, prod_schedule, daily_metrics, inventory_batches = read_masterpapi(MASTER_PAPI_PATH)

    print(f"\n[2/3] Reading MarginPapi (all SKUs) + Inventory...")
    margin_skus = read_all_margin_skus()
    inventory_snap = read_inventory()

    print(f"\n[3/3] Upserting to Supabase...")
    upsert_supabase(monthly, weekly, prod_runs, prod_efficiency, prod_schedule, margin_skus, inventory_snap, daily_metrics, inventory_batches)

    print("\n✓ Sync complete. Refresh the dashboard to see updated data.")
    print("  https://ceo-dashboard.vercel.app")

if __name__ == '__main__':
    main()
