"""
sync_marginpapi.py
------------------
Reads all active MarginPapi Excel files and syncs costing data to Supabase.

Tables written:
  costing_cogs         — COGS breakdown per SKU/variant (ingredients + packaging + OH)
  costing_ingredients  — Ingredient list per SKU with quantities and costs
  costing_packaging    — Packaging items per SKU/variant with unit cost
  costing_overheads    — OH assumptions per SKU (hourly rate, super %, rent)

Usage:
    python scripts/sync_marginpapi.py
"""

import os
import re
import sys
import shutil
import tempfile
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
import openpyxl

SCRIPT_DIR = Path(__file__).parent
load_dotenv(SCRIPT_DIR.parent / '.env.secret')

SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_KEY')

HOME = Path.home()
ONEDRIVE_MAC = HOME / "Library/CloudStorage/OneDrive-umamipapi.com.au"
ONEDRIVE_WIN = HOME / "OneDrive - umamipapi.com.au"
ONEDRIVE = ONEDRIVE_MAC if ONEDRIVE_MAC.exists() else ONEDRIVE_WIN

COSTING_DIR = ONEDRIVE / "DatabasePapi" / "Costing"

MARGINPAPI_FILES = [
    "MarginPapi - Chilli Oil (Apr 26).xlsx",
    "MarginPapi - Hot Honey (Apr 26).xlsx",
    "MarginPapi - Nandos PSE (Apr 26).xlsx",
    "MarginPapi - Nandos PCRK (Apr 26).xlsx",
    "MarginPapi - Mayo (Apr26).xlsx",
]

SKU_NAMES = {
    "MarginPapi - Chilli Oil (Apr 26).xlsx":  "Chilli Oil",
    "MarginPapi - Hot Honey (Apr 26).xlsx":   "Hot Honey",
    "MarginPapi - Nandos PSE (Apr 26).xlsx":  "Nandos PSE",
    "MarginPapi - Nandos PCRK (Apr 26).xlsx": "Nandos PCRK",
    "MarginPapi - Mayo (Apr26).xlsx":          "Mayo",
}


def open_wb(path: Path):
    tmp = tempfile.mktemp(suffix='.xlsx')
    shutil.copy2(str(path), tmp)
    wb = openpyxl.load_workbook(tmp, data_only=True)
    os.unlink(tmp)
    return wb


def _float(v):
    try:
        f = float(v)
        return None if f == 0 else round(f, 6)
    except (TypeError, ValueError):
        return None


def _parse_date(s):
    """Parse expiry date from batch code like OG-LRG-14.10.27 → 2027-10-14."""
    if not s or not isinstance(s, str):
        return None
    m = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})$', s)
    if not m:
        return None
    d, mo, y = m.groups()
    y = int(y)
    if y < 100:
        y += 2000
    try:
        return f"{y:04d}-{int(mo):02d}-{int(d):02d}"
    except Exception:
        return None


# ── Readers ──────────────────────────────────────────────────────────────────

def read_ingredients(wb, sku_name: str, source_file: str) -> list[dict]:
    ws = wb['Ingredients']
    records = []
    # Headers at row 14: col E=Component, G=qty_per_unit, H=qty_per_batch, I=unit_cost, J=batch_cost
    # Data starts at row 16
    now = datetime.utcnow().isoformat()
    for row in ws.iter_rows(min_row=16, max_row=ws.max_row, values_only=True):
        component = row[4] if len(row) > 4 else None  # col E = index 4
        if not component or not isinstance(component, str):
            continue
        component = component.strip()
        if not component:
            continue
        qty_per_unit   = _float(row[6])  if len(row) > 6  else None  # col G
        qty_per_batch  = _float(row[7])  if len(row) > 7  else None  # col H
        unit_cost      = _float(row[8])  if len(row) > 8  else None  # col I
        batch_cost_raw = row[9]          if len(row) > 9  else None  # col J
        batch_cost     = _float(batch_cost_raw)
        records.append({
            'sku_name':      sku_name,
            'source_file':   source_file,
            'component':     component,
            'qty_per_unit':  qty_per_unit,
            'qty_per_batch': qty_per_batch,
            'unit_cost':     unit_cost,
            'batch_cost':    batch_cost,
            'synced_at':     now,
        })
    return records


def read_packaging(wb, sku_name: str, source_file: str) -> list[dict]:
    ws = wb['Packaging']
    records = []
    now = datetime.utcnow().isoformat()
    # Layout: col E=packaging_type, col F=product_variant, col H=unit_cost (first period)
    current_type = None
    for row in ws.iter_rows(min_row=10, max_row=ws.max_row, values_only=True):
        # col C (index 2) = section number flag
        if row[2] is not None and isinstance(row[2], (int, float)):
            # section header — next few rows will set type
            pass
        pkg_type   = row[4] if len(row) > 4 else None  # col E
        variant    = row[5] if len(row) > 5 else None  # col F
        unit_cost  = _float(row[7]) if len(row) > 7 else None  # col H = first period

        if pkg_type and isinstance(pkg_type, str) and pkg_type.strip():
            current_type = pkg_type.strip()
        if not variant or not isinstance(variant, str):
            continue
        variant = variant.strip()
        if not variant:
            continue
        if unit_cost is None:
            continue
        records.append({
            'sku_name':       sku_name,
            'source_file':    source_file,
            'packaging_type': current_type or 'Unknown',
            'product_variant': variant,
            'unit_cost':      unit_cost,
            'synced_at':      now,
        })
    return records


def read_overheads(wb, sku_name: str, source_file: str) -> dict:
    ws = wb['OH']
    now = datetime.utcnow().isoformat()
    # Core assumptions: row 6=hourly_rate (col G=index 6), row 7=super%, row 8=rent
    rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
    hourly_rate = None
    super_pct   = None
    annual_rent = None
    for r in rows:
        label = r[5] if len(r) > 5 else None  # col F
        if not label:
            continue
        val = r[6] if len(r) > 6 else None  # col G
        if isinstance(label, str):
            l = label.strip().lower()
            if 'hourly rate' in l or 'average hourly' in l:
                hourly_rate = _float(val)
            elif 'super' in l:
                super_pct = _float(val)
            elif 'rent' in l:
                annual_rent = _float(val)
    return {
        'sku_name':    sku_name,
        'source_file': source_file,
        'hourly_rate': hourly_rate,
        'super_pct':   super_pct,
        'annual_rent': annual_rent,
        'synced_at':   now,
    }


def read_cogs(wb, sku_name: str, source_file: str) -> list[dict]:
    ws = wb['COGS']
    now = datetime.utcnow().isoformat()
    records = []
    # Structure per product block:
    #   Row N:   col C=1 (flag), col F=product_variant
    #   Row N+3: col F='Staff', col H=staff_hours (not needed)
    #   Row N+4: col F='Batches', col H=batches
    #   Row N+5: col F='Units per batch', col H=units_per_batch
    #   Row N+6: col F='Total units', col H=total_units
    #   Row N+8: col E=variant, col F='Ingredients', col H=ingredients_cogs
    #   Row N+9: col E=variant, col F='Packaging', col H=packaging_cogs
    #   Row N+10:col E=variant, col F='Overheads', col H=overheads_cogs
    #   Row N+12:col F='Total COGS', col H=total_cogs
    # col H = index 7

    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    i = 0
    while i < len(all_rows):
        r = all_rows[i]
        # Detect product block: col C (index 2) == 1 AND col F (index 5) has a non-empty string
        if r[2] == 1 and len(r) > 5 and isinstance(r[5], str) and r[5].strip() and r[5].strip() not in ('COGS constants', 'COGS by product', 'Ingredients', 'Packaging', 'Overheads', 'Total COGS', 'Staff', 'Batches', 'Units per batch', 'Total units'):
            variant = r[5].strip()
            batches = None
            units_per_batch = None
            total_units = None
            ingredients_cogs = None
            packaging_cogs = None
            overheads_cogs = None
            total_cogs = None
            # Scan next 15 rows for this product's data
            for j in range(i + 1, min(i + 16, len(all_rows))):
                rj = all_rows[j]
                label = rj[5] if len(rj) > 5 else None
                val   = rj[7] if len(rj) > 7 else None  # col H = first data period
                if not isinstance(label, str):
                    continue
                l = label.strip()
                if l == 'Batches':
                    batches = _float(val)
                elif l == 'Units per batch':
                    units_per_batch = _float(val)
                elif l == 'Total units':
                    total_units = _float(val)
                elif l == 'Ingredients':
                    ingredients_cogs = _float(val)
                elif l == 'Packaging':
                    packaging_cogs = _float(val)
                elif l == 'Overheads':
                    overheads_cogs = _float(val)
                elif l == 'Total COGS':
                    total_cogs = _float(val)
                    break
            records.append({
                'sku_name':         sku_name,
                'source_file':      source_file,
                'product_variant':  variant,
                'ingredients_cogs': ingredients_cogs,
                'packaging_cogs':   packaging_cogs,
                'overheads_cogs':   overheads_cogs,
                'total_cogs':       total_cogs,
                'batches_per_month': batches,
                'units_per_batch':  units_per_batch,
                'total_units':      total_units,
                'synced_at':        now,
            })
        i += 1
    return records


# ── Supabase sync ─────────────────────────────────────────────────────────────

def get_supabase():
    from supabase import create_client
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("ERROR: Missing SUPABASE_URL or SUPABASE_SERVICE_KEY in .env.secret")
        sys.exit(1)
    return create_client(SUPABASE_URL, SUPABASE_KEY)


def upsert_table(sb, table: str, records: list, conflict_cols: str):
    if not records:
        print(f"  {table}: no records to upsert")
        return
    BATCH = 200
    total = 0
    for i in range(0, len(records), BATCH):
        chunk = records[i:i+BATCH]
        sb.table(table).upsert(chunk, on_conflict=conflict_cols).execute()
        total += len(chunk)
    print(f"  {table}: {total} rows upserted")


def wipe_sku(sb, table: str, sku_name: str):
    sb.table(table).delete().eq('sku_name', sku_name).execute()


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    sb = get_supabase()

    all_cogs         = []
    all_ingredients  = []
    all_packaging    = []
    all_overheads    = []

    for fname in MARGINPAPI_FILES:
        path = COSTING_DIR / fname
        if not path.exists():
            print(f"  SKIP (not found): {fname}")
            continue

        sku_name    = SKU_NAMES[fname]
        source_file = fname
        print(f"\n→ {sku_name} ({fname})")

        try:
            wb = open_wb(path)
        except Exception as e:
            print(f"  ERROR opening: {e}")
            continue

        try:
            ingr = read_ingredients(wb, sku_name, source_file)
            print(f"  ingredients: {len(ingr)} rows")
            all_ingredients.extend(ingr)
        except Exception as e:
            print(f"  WARN ingredients: {e}")

        try:
            pkg = read_packaging(wb, sku_name, source_file)
            print(f"  packaging: {len(pkg)} rows")
            all_packaging.extend(pkg)
        except Exception as e:
            print(f"  WARN packaging: {e}")

        try:
            oh = read_overheads(wb, sku_name, source_file)
            print(f"  overheads: hourly_rate={oh['hourly_rate']}, super={oh['super_pct']}, rent={oh['annual_rent']}")
            all_overheads.append(oh)
        except Exception as e:
            print(f"  WARN overheads: {e}")

        try:
            cogs = read_cogs(wb, sku_name, source_file)
            print(f"  cogs: {len(cogs)} variants")
            all_cogs.extend(cogs)
        except Exception as e:
            print(f"  WARN cogs: {e}")

    print("\n── Syncing to Supabase ──")
    upsert_table(sb, 'costing_cogs',         all_cogs,        'sku_name,product_variant')
    upsert_table(sb, 'costing_ingredients',  all_ingredients, 'sku_name,component')
    upsert_table(sb, 'costing_packaging',    all_packaging,   'sku_name,packaging_type,product_variant')

    # Overheads: upsert by sku_name
    if all_overheads:
        sb.table('costing_overheads').upsert(all_overheads, on_conflict='sku_name').execute()
        print(f"  costing_overheads: {len(all_overheads)} rows upserted")

    print("\nDone.")


if __name__ == '__main__':
    main()
